import csv
import io
import json
import re

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.views import View
from django.views.generic import CreateView, UpdateView, DetailView, DeleteView

from apps.users.models import User
from utils.dynamic_filter import apply_dynamic_filters, fields_for_js, _contacto_fields
from .forms import (
    ContactoForm, ContactoImportForm, EtapaChangeForm, CampoPersonalizadoForm, TipoContactoForm,
)
from django.utils import timezone

from .models import (
    Contacto, HistorialEtapa, Plan, CampoPersonalizado, Documento, TipoContacto, PipelineStage,
    NotaContacto, Etiqueta, FiltroGuardado,
    PRIORIDAD_CHOICES, ORIGEN_CHOICES,
)

# Columns that are recognized as standard Contacto fields (case-insensitive)
_KNOWN_COLUMNS = {
    'nombre_completo', 'nombre', 'name', 'full_name', 'apellido',
    'telefono', 'phone', 'tel', 'celular', 'movil',
    'codigo_pais', 'codigopais', 'country_code', 'cod_pais',
    'email', 'correo', 'mail',
    'dni', 'documento', 'cedula', 'rut',
    'localidad', 'ciudad', 'city',
    'provincia', 'province', 'region',
    'tipo', 'tipo_contacto',
    'prioridad', 'priority',
    'notas', 'notes', 'observaciones', 'comentarios',
    'plan', 'plan_interes',
    'origen', 'origin', 'source', 'fuente',
    'grupo_familiar', 'grupo',
    'agente', 'agent', 'vendedor', 'asesor',
}

_PRIORIDAD_MAP = {
    'alta': 'alta', 'high': 'alta',
    'media': 'media', 'medium': 'media', 'normal': 'media',
    'baja': 'baja', 'low': 'baja',
}

_ORIGEN_MAP = {
    'web': 'web', 'campaña': 'campana', 'campana': 'campana',
    'referido': 'referido', 'llamada': 'llamada', 'whatsapp': 'whatsapp',
}


def _read_file(archivo) -> tuple:
    """Read CSV or Excel. Returns (headers: list[str], rows: list[dict])."""
    name = archivo.name.lower()
    if name.endswith('.xlsx') or name.endswith('.xls'):
        import openpyxl
        wb = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            return [], []
        headers = [str(h).strip() if h is not None else f'col_{i}' for i, h in enumerate(all_rows[0])]
        rows = []
        for raw in all_rows[1:]:
            if all(v is None or str(v).strip() == '' for v in raw):
                continue
            rows.append({headers[i]: (str(raw[i]).strip() if raw[i] is not None else '') for i in range(len(headers))})
        return headers, rows
    else:
        try:
            content = archivo.read().decode('utf-8-sig')
        except UnicodeDecodeError:
            archivo.seek(0)
            content = archivo.read().decode('latin-1')
        reader = csv.DictReader(io.StringIO(content))
        headers = list(reader.fieldnames or [])
        rows = [dict(r) for r in reader if any(v.strip() for v in r.values())]
        return headers, rows


def _normalize_phone(raw: str, codigo_pais: str = '54') -> str:
    """Return a normalized phone like +54XXXXXXXXXX, or '' if invalid."""
    cleaned = re.sub(r'[^\d+]', '', str(raw)).strip()
    if not cleaned:
        return ''
    if cleaned.startswith('+'):
        digits = re.sub(r'\D', '', cleaned)
    else:
        digits_stripped = cleaned.lstrip('0') or cleaned
        codigo = re.sub(r'\D', '', str(codigo_pais)).lstrip('0') or '54'
        digits = codigo + digits_stripped
    if len(digits) < 7:
        return ''
    return '+' + digits


def _get(row_lower: dict, *keys, default='') -> str:
    for k in keys:
        v = row_lower.get(k.lower(), '')
        if v and str(v).strip():
            return str(v).strip()
    return default


def _resolve_agente(nombre_agente: str, agentes_cache: dict) -> object:
    """Look up an agent by full name or email, case-insensitive. Returns User or None."""
    if not nombre_agente:
        return None
    key = nombre_agente.strip().lower()
    if key in agentes_cache:
        return agentes_cache[key]
    if not agentes_cache:
        for u in User.objects.filter(is_active=True):
            agentes_cache[u.get_full_name().lower()] = u
            agentes_cache[u.username.lower()] = u
            agentes_cache[u.email.lower()] = u
    return agentes_cache.get(key)


def _process_row(row: dict, plans_by_name: dict, tipos_by_slug: dict, default_tipo,
                 actualizar: bool, default_agente, agentes_cache: dict | None = None) -> tuple:
    """Returns ('created'|'updated'|'skipped'|'error', msg, datos_extra_keys)."""
    if agentes_cache is None:
        agentes_cache = {}
    row_lower = {k.lower().strip(): v for k, v in row.items()}

    nombre = _get(row_lower, 'nombre_completo', 'nombre', 'name', 'full_name', 'apellido')
    if not nombre:
        return 'error', 'Nombre vacío', []

    raw_phone = _get(row_lower, 'telefono', 'phone', 'tel', 'celular', 'movil')
    codigo_pais = _get(row_lower, 'codigo_pais', 'codigopais', 'country_code', 'cod_pais', default='54')
    phone = _normalize_phone(raw_phone, codigo_pais)
    if not phone:
        return 'error', f'Teléfono inválido: "{raw_phone}"', []

    email = _get(row_lower, 'email', 'correo', 'mail')
    raw_dni = re.sub(r'\D', '', _get(row_lower, 'dni', 'documento', 'cedula', 'rut'))
    dni = raw_dni[:8] if raw_dni else ''
    if dni and len(dni) < 7:
        dni = dni.zfill(7)

    localidad = _get(row_lower, 'localidad', 'ciudad', 'city')
    provincia = _get(row_lower, 'provincia', 'province', 'region')
    notas = _get(row_lower, 'notas', 'notes', 'observaciones', 'comentarios')
    plan_nombre = _get(row_lower, 'plan', 'plan_interes')

    origen_raw = _get(row_lower, 'origen', 'origin', 'source', 'fuente').lower()
    origen = _ORIGEN_MAP.get(origen_raw, 'campana')

    plan = plans_by_name.get(plan_nombre.lower()) if plan_nombre else None

    prioridad_raw = _get(row_lower, 'prioridad', 'priority').lower()
    prioridad = _PRIORIDAD_MAP.get(prioridad_raw, 'media')

    tipo_raw = _get(row_lower, 'tipo', 'tipo_contacto').lower()
    tipo = tipos_by_slug.get(tipo_raw) or default_tipo

    agente_raw = _get(row_lower, 'agente', 'agent', 'vendedor', 'asesor')
    agente = _resolve_agente(agente_raw, agentes_cache) if agente_raw else default_agente

    datos_extra = {
        k.strip(): str(v).strip()
        for k, v in row.items()
        if k.lower().strip() not in _KNOWN_COLUMNS and v and str(v).strip()
    }
    extra_keys = list(datos_extra.keys())

    existing = Contacto.objects.filter(telefono=phone).first()
    if not existing and dni:
        existing = Contacto.objects.filter(dni=dni).first()

    if existing:
        if not actualizar:
            return 'skipped', '', extra_keys
        updated = []
        if not existing.nombre_completo and nombre:
            existing.nombre_completo = nombre; updated.append('nombre_completo')
        if not existing.email and email:
            existing.email = email; updated.append('email')
        if not existing.localidad and localidad:
            existing.localidad = localidad; updated.append('localidad')
        if not existing.provincia and provincia:
            existing.provincia = provincia; updated.append('provincia')
        if not existing.notas and notas:
            existing.notas = notas; updated.append('notas')
        if datos_extra:
            current = existing.datos_extra or {}
            current.update(datos_extra)
            existing.datos_extra = current
            updated.append('datos_extra')
        if updated:
            existing.save(update_fields=updated + ['updated_at'])
            return 'updated', '', extra_keys
        return 'skipped', '', extra_keys

    stage = tipo.pipeline.stages.order_by('orden').first() if tipo.pipeline_id else None
    contacto = Contacto.objects.create(
        tipo=tipo,
        nombre_completo=nombre,
        dni=dni,
        telefono=phone,
        email=email,
        localidad=localidad,
        provincia=provincia,
        notas=notas,
        origen=origen,
        plan_interes=plan,
        stage=stage,
        prioridad=prioridad,
        agente=agente,
        datos_extra=datos_extra,
    )
    HistorialEtapa.objects.create(contacto=contacto, etapa_nueva=stage, nota='Importado.')
    return 'created', '', extra_keys


class ContactoQuerysetMixin:
    def get_base_queryset(self):
        qs = (Contacto.objects
              .select_related('agente', 'plan_interes', 'tipo', 'stage', 'stage__pipeline')
              .prefetch_related('conversacion_whatsapp'))
        if not self.request.user.can_see_all_leads:
            qs = qs.filter(agente=self.request.user)
        return qs


class ContactoListView(LoginRequiredMixin, ContactoQuerysetMixin, View):
    template_name = 'contactos/contacto_list.html'

    def get(self, request):
        qs = self.get_base_queryset()

        q = request.GET.get('q', '').strip()
        if q:
            qs = qs.filter(
                Q(nombre_completo__icontains=q) | Q(dni__icontains=q) | Q(telefono__icontains=q)
            )

        tipo_slug = request.GET.get('tipo', '').strip()
        tipo_actual = None
        if tipo_slug:
            tipo_actual = TipoContacto.objects.filter(slug=tipo_slug).first()
            if tipo_actual:
                qs = qs.filter(tipo=tipo_actual)

        etiqueta_slug = request.GET.get('etiqueta', '').strip()
        if etiqueta_slug:
            qs = qs.filter(etiquetas__slug=etiqueta_slug)

        field_defs = _contacto_fields(tipo_actual)
        qs = apply_dynamic_filters(qs, request, field_defs)

        paginator = Paginator(qs, 25)
        page = paginator.get_page(request.GET.get('page'))

        agents = User.objects.filter(is_active=True).order_by('first_name') if request.user.can_see_all_leads else None

        active_filters = list(zip(
            request.GET.getlist('fc'),
            request.GET.getlist('fo'),
            request.GET.getlist('fv'),
        ))

        filtros_guardados = FiltroGuardado.objects.filter(usuario=request.user)
        if tipo_actual:
            filtros_guardados = filtros_guardados.filter(
                Q(tipo_contacto=tipo_actual) | Q(tipo_contacto__isnull=True)
            )

        return render(request, self.template_name, {
            'contactos':       page,
            'tipos':           TipoContacto.objects.filter(activo=True),
            'tipo_actual':     tipo_actual,
            'agents':          agents,
            'total_count':     qs.count(),
            'q':               q,
            'active_filters':  active_filters,
            'filter_fields':   json.dumps(fields_for_js(field_defs, request.user), ensure_ascii=False),
            'filtros_guardados': filtros_guardados,
            'etiquetas':       Etiqueta.objects.all(),
        })


class ContactoKanbanView(LoginRequiredMixin, ContactoQuerysetMixin, View):
    template_name = 'contactos/contacto_kanban.html'

    def get(self, request):
        tipos = list(TipoContacto.objects.filter(activo=True, pipeline__isnull=False))
        if not tipos:
            return render(request, self.template_name, {'tipos': [], 'tipo_actual': None, 'columns': {}})

        tipo_slug = request.GET.get('tipo', '').strip()
        tipo_actual = next((t for t in tipos if t.slug == tipo_slug), None) or tipos[0]

        qs = self.get_base_queryset().filter(tipo=tipo_actual)
        columns = {}
        for stage in tipo_actual.pipeline.stages.all():
            columns[stage.pk] = {
                'stage': stage,
                'contactos': qs.filter(stage=stage).order_by('-prioridad', '-updated_at')[:50],
            }
        return render(request, self.template_name, {
            'tipos': tipos,
            'tipo_actual': tipo_actual,
            'columns': columns,
        })


class ContactoCreateView(LoginRequiredMixin, CreateView):
    model = Contacto
    form_class = ContactoForm
    template_name = 'contactos/contacto_form.html'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        tipo_slug = self.request.GET.get('tipo', '')
        if tipo_slug and not self.request.POST:
            tipo = TipoContacto.objects.filter(slug=tipo_slug, activo=True).first()
            if tipo:
                kwargs['initial'] = {**kwargs.get('initial', {}), 'tipo': tipo.pk}
        return kwargs

    def form_valid(self, form):
        contacto = form.save(commit=False)
        if not contacto.agente and not self.request.user.can_see_all_leads:
            contacto.agente = self.request.user
        if not contacto.stage_id and contacto.tipo.pipeline_id:
            contacto.stage = contacto.tipo.pipeline.stages.order_by('orden').first()
        contacto.stage_updated_at = timezone.now()
        contacto.save()
        HistorialEtapa.objects.create(
            contacto=contacto,
            etapa_nueva=contacto.stage,
            cambiado_por=self.request.user,
            nota='Contacto creado.',
        )
        messages.success(self.request, 'Contacto creado correctamente.')
        return redirect('contactos:detail', pk=contacto.pk)


class ContactoDetailView(LoginRequiredMixin, ContactoQuerysetMixin, DetailView):
    template_name = 'contactos/contacto_detail.html'
    context_object_name = 'contacto'

    def get_object(self):
        return get_object_or_404(self.get_base_queryset(), pk=self.kwargs['pk'])

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        contacto = self.object
        ctx['historial'] = contacto.historial_etapas.select_related('cambiado_por', 'etapa_anterior', 'etapa_nueva').all()
        ctx['etapa_form'] = EtapaChangeForm(contacto=contacto, initial={'etapa': contacto.stage_id})
        ctx['tareas'] = contacto.tareas.select_related('agente').order_by('-fecha_programada')[:10]
        ctx['cotizaciones'] = contacto.cotizaciones.order_by('-created_at')[:5]
        ctx['mensajes'] = contacto.mensajes_whatsapp.order_by('-timestamp')[:20]

        campos = list(CampoPersonalizado.objects.prefetch_related('reglas').filter(activo=True).filter(
            Q(tipo_contacto__isnull=True) | Q(tipo_contacto=contacto.tipo)
        ))
        ctx['campos'] = campos
        ctx['campos_config'] = {
            c.slug: {
                'obligatorio': c.requerido or any(
                    r.accion == 'obligatorio' and r.evaluar(contacto) for r in c.reglas.all()
                ),
                'oculto': any(r.accion == 'oculto' and r.evaluar(contacto) for r in c.reglas.all()),
                'solo_lectura': any(r.accion == 'solo_lectura' and r.evaluar(contacto) for r in c.reglas.all()),
            }
            for c in campos
        }
        ctx['campos_config_json'] = json.dumps(ctx['campos_config'])
        ctx['documentos'] = contacto.documentos.select_related('subido_por').all()
        ctx['documento_tipos'] = Documento.TIPO_CHOICES
        ctx['otros_tipos'] = TipoContacto.objects.filter(activo=True).exclude(pk=contacto.tipo_id)

        # Notas internas
        ctx['notas_internas'] = contacto.notas_internas.select_related('autor').prefetch_related('menciones').all()

        # Etiquetas
        ctx['etiquetas_contacto'] = list(contacto.etiquetas.values_list('slug', flat=True))
        ctx['todas_etiquetas'] = Etiqueta.objects.all()

        # Leads frios: días sin cambio de etapa
        ctx['dias_en_etapa'] = None
        if contacto.stage_updated_at:
            ctx['dias_en_etapa'] = (timezone.now() - contacto.stage_updated_at).days

        # Agentes disponibles para asignación
        ctx['agentes_disponibles'] = User.objects.filter(is_active=True, disponible=True).order_by('first_name')

        # Planes
        ctx['planes'] = Plan.objects.filter(activo=True)

        # Stage gate: campos requeridos para cada etapa del pipeline
        stage_gates = {}
        if contacto.tipo.pipeline_id:
            for stage in contacto.tipo.pipeline.stages.all():
                requeridos = []
                for campo in campos:
                    for regla in campo.reglas.all():
                        if (regla.accion == 'obligatorio'
                                and regla.condicion_tipo == 'etapa'
                                and regla.condicion_etapa_id == stage.pk):
                            val = (contacto.datos_extra or {}).get(campo.slug)
                            if not val:
                                requeridos.append(campo.nombre)
                if requeridos:
                    stage_gates[stage.pk] = requeridos
        ctx['stage_gates_json'] = json.dumps(stage_gates)

        # Unified activity timeline
        tl = []
        for h in contacto.historial_etapas.select_related('cambiado_por', 'etapa_anterior', 'etapa_nueva').all():
            ant = h.etapa_anterior.nombre if h.etapa_anterior else '—'
            nvo = h.etapa_nueva.nombre if h.etapa_nueva else '—'
            tl.append({
                'fecha': h.fecha,
                'icon': 'arrow-right-circle',
                'color': 'primary',
                'texto': f'Etapa: {ant} → {nvo}',
                'autor': h.cambiado_por.display_name if h.cambiado_por else 'Sistema',
                'nota': h.nota or '',
            })
        for n in contacto.notas_internas.select_related('autor').all():
            tl.append({
                'fecha': n.created_at,
                'icon': 'chat-left-text',
                'color': 'success',
                'texto': f'Nota de {n.autor.display_name if n.autor else "Sistema"}',
                'autor': n.autor.display_name if n.autor else 'Sistema',
                'nota': n.texto[:200],
            })
        for t in contacto.tareas.select_related('agente').all():
            tl.append({
                'fecha': t.fecha_programada,
                'icon': 'check2-square',
                'color': 'warning',
                'texto': f'Tarea: {t.get_tipo_display()} ({t.get_status_display()})',
                'autor': t.agente.display_name if t.agente else 'Sin asignar',
                'nota': t.descripcion[:100] if t.descripcion else '',
            })
        ctx['timeline'] = sorted(tl, key=lambda x: x['fecha'], reverse=True)

        return ctx


class ContactoUpdateView(LoginRequiredMixin, ContactoQuerysetMixin, UpdateView):
    model = Contacto
    form_class = ContactoForm
    template_name = 'contactos/contacto_form.html'

    def get_object(self):
        return get_object_or_404(self.get_base_queryset(), pk=self.kwargs['pk'])

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        form.save()
        messages.success(self.request, 'Contacto actualizado correctamente.')
        return redirect('contactos:detail', pk=self.object.pk)


class ContactoDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = Contacto
    template_name = 'contactos/contacto_confirm_delete.html'
    success_url = reverse_lazy('contactos:list')

    def test_func(self):
        return self.request.user.can_see_all_leads

    def form_valid(self, form):
        messages.success(self.request, 'Contacto eliminado.')
        return super().form_valid(form)


class ContactoUpdateCamposView(LoginRequiredMixin, ContactoQuerysetMixin, View):
    """Save custom field values for a contacto."""

    def post(self, request, pk):
        contacto = get_object_or_404(self.get_base_queryset(), pk=pk)
        campos = list(CampoPersonalizado.objects.prefetch_related('reglas').filter(activo=True).filter(
            Q(tipo_contacto__isnull=True) | Q(tipo_contacto=contacto.tipo)
        ))
        extra = dict(contacto.datos_extra or {})
        errors = []

        for campo in campos:
            oculto = any(r.accion == 'oculto' and r.evaluar(contacto) for r in campo.reglas.all())
            if oculto:
                continue

            if campo.tipo == CampoPersonalizado.TIPO_BOOLEANO:
                extra[campo.slug] = bool(request.POST.get(f'campo_{campo.slug}'))
            else:
                val = request.POST.get(f'campo_{campo.slug}', '').strip()
                obligatorio = campo.requerido or any(
                    r.accion == 'obligatorio' and r.evaluar(contacto) for r in campo.reglas.all()
                )
                if obligatorio and not val:
                    errors.append(f'"{campo.nombre}" es obligatorio.')
                if val:
                    extra[campo.slug] = val
                else:
                    extra.pop(campo.slug, None)

        if errors:
            for e in errors:
                messages.error(request, e)
            return redirect('contactos:detail', pk=pk)

        contacto.datos_extra = extra
        contacto.save(update_fields=['datos_extra'])
        messages.success(request, 'Campos guardados.')
        return redirect('contactos:detail', pk=pk)


class ContactoCambiarTipoView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Change a contacto's tipo (e.g. Lead → Cliente) and reset its stage to the new pipeline's first stage."""

    def test_func(self):
        return self.request.user.can_see_all_leads

    def post(self, request, pk):
        contacto = get_object_or_404(Contacto, pk=pk)
        tipo_id = request.POST.get('tipo_id', '').strip()
        nuevo_tipo = get_object_or_404(TipoContacto, pk=tipo_id, activo=True)
        if nuevo_tipo.pk == contacto.tipo_id:
            messages.info(request, f'{contacto.nombre_completo} ya es de tipo {nuevo_tipo.nombre}.')
            return redirect('contactos:detail', pk=pk)

        tipo_anterior = contacto.tipo
        contacto.tipo = nuevo_tipo
        contacto.stage = nuevo_tipo.pipeline.stages.order_by('orden').first() if nuevo_tipo.pipeline_id else None
        contacto.save(update_fields=['tipo', 'stage', 'updated_at'])
        HistorialEtapa.objects.create(
            contacto=contacto,
            etapa_nueva=contacto.stage,
            cambiado_por=request.user,
            nota=f'Tipo cambiado de "{tipo_anterior.nombre}" a "{nuevo_tipo.nombre}".',
        )
        messages.success(request, f'{contacto.nombre_completo} ahora es {nuevo_tipo.nombre}.')
        return redirect('contactos:detail', pk=pk)


# ── Documentos ──────────────────────────────────────────

class DocumentoUploadView(LoginRequiredMixin, View):
    def post(self, request, contacto_pk):
        contacto = get_object_or_404(Contacto, pk=contacto_pk)
        if not request.user.can_see_all_leads and contacto.agente != request.user:
            return HttpResponseForbidden()
        archivo = request.FILES.get('archivo')
        nombre = request.POST.get('nombre', '').strip()
        tipo = request.POST.get('tipo', Documento.TIPO_OTRO)
        if not archivo:
            messages.error(request, 'Seleccioná un archivo.')
            return redirect('contactos:detail', pk=contacto_pk)
        if archivo.size > 20 * 1024 * 1024:
            messages.error(request, 'El archivo no puede superar 20 MB.')
            return redirect('contactos:detail', pk=contacto_pk)
        Documento.objects.create(
            contacto=contacto,
            nombre=nombre or archivo.name,
            tipo=tipo,
            archivo=archivo,
            subido_por=request.user,
        )
        messages.success(request, 'Documento subido correctamente.')
        return redirect('contactos:detail', pk=contacto_pk)


class DocumentoDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        doc = get_object_or_404(Documento, pk=pk)
        contacto_pk = doc.contacto_id
        if not request.user.can_see_all_leads and doc.contacto.agente != request.user:
            return HttpResponseForbidden()
        doc.archivo.delete(save=False)
        doc.delete()
        messages.success(request, 'Documento eliminado.')
        return redirect('contactos:detail', pk=contacto_pk)


# ── Tipos de contacto CRUD (configurable desde el panel) ──

class SupervisorRequiredMixin(UserPassesTestMixin):
    def test_func(self):
        return self.request.user.can_see_all_leads


class TipoContactoListView(LoginRequiredMixin, SupervisorRequiredMixin, View):
    template_name = 'contactos/tipo_list.html'

    def get(self, request):
        tipos = TipoContacto.objects.select_related('pipeline').all()
        return render(request, self.template_name, {'tipos': tipos})


class TipoContactoCreateView(LoginRequiredMixin, SupervisorRequiredMixin, View):
    template_name = 'contactos/tipo_form.html'

    def get(self, request):
        return render(request, self.template_name, {'form': TipoContactoForm()})

    def post(self, request):
        form = TipoContactoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Tipo de contacto creado.')
            return redirect('contactos:tipos')
        return render(request, self.template_name, {'form': form})


class TipoContactoUpdateView(LoginRequiredMixin, SupervisorRequiredMixin, View):
    template_name = 'contactos/tipo_form.html'

    def get(self, request, pk):
        tipo = get_object_or_404(TipoContacto, pk=pk)
        return render(request, self.template_name, {'form': TipoContactoForm(instance=tipo), 'tipo': tipo})

    def post(self, request, pk):
        tipo = get_object_or_404(TipoContacto, pk=pk)
        form = TipoContactoForm(request.POST, instance=tipo)
        if form.is_valid():
            form.save()
            messages.success(request, 'Tipo de contacto actualizado.')
            return redirect('contactos:tipos')
        return render(request, self.template_name, {'form': form, 'tipo': tipo})


class TipoContactoDeleteView(LoginRequiredMixin, SupervisorRequiredMixin, View):
    def post(self, request, pk):
        tipo = get_object_or_404(TipoContacto, pk=pk)
        if tipo.es_sistema:
            messages.error(request, f'"{tipo.nombre}" es un tipo del sistema y no se puede eliminar.')
            return redirect('contactos:tipos')
        if tipo.contactos.exists():
            messages.error(request, f'No se puede eliminar "{tipo.nombre}": tiene contactos asociados. Reasigná o eliminá esos contactos primero.')
            return redirect('contactos:tipos')
        tipo.delete()
        messages.success(request, 'Tipo de contacto eliminado.')
        return redirect('contactos:tipos')


# ── Campos personalizados CRUD ────────────────────────────

class CampoListView(LoginRequiredMixin, SupervisorRequiredMixin, View):
    template_name = 'contactos/campo_list.html'

    def get(self, request):
        campos = CampoPersonalizado.objects.select_related('tipo_contacto').all()
        return render(request, self.template_name, {'campos': campos})


class CampoCreateView(LoginRequiredMixin, SupervisorRequiredMixin, View):
    template_name = 'contactos/campo_form.html'

    def get(self, request):
        return render(request, self.template_name, {'form': CampoPersonalizadoForm()})

    def post(self, request):
        form = CampoPersonalizadoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Campo creado.')
            return redirect('contactos:campos')
        return render(request, self.template_name, {'form': form})


class CampoUpdateView(LoginRequiredMixin, SupervisorRequiredMixin, View):
    template_name = 'contactos/campo_form.html'

    def get(self, request, pk):
        campo = get_object_or_404(CampoPersonalizado, pk=pk)
        return render(request, self.template_name, {'form': CampoPersonalizadoForm(instance=campo), 'campo': campo})

    def post(self, request, pk):
        campo = get_object_or_404(CampoPersonalizado, pk=pk)
        form = CampoPersonalizadoForm(request.POST, instance=campo)
        if form.is_valid():
            form.save()
            messages.success(request, 'Campo actualizado.')
            return redirect('contactos:campos')
        return render(request, self.template_name, {'form': form, 'campo': campo})


class CampoDeleteView(LoginRequiredMixin, SupervisorRequiredMixin, DeleteView):
    model = CampoPersonalizado
    template_name = 'contactos/campo_confirm_delete.html'
    success_url = reverse_lazy('contactos:campos')


class CampoToggleView(LoginRequiredMixin, SupervisorRequiredMixin, View):
    def post(self, request, pk):
        campo = get_object_or_404(CampoPersonalizado, pk=pk)
        campo.activo = not campo.activo
        campo.save(update_fields=['activo'])
        return redirect('contactos:campos')


class ContactoEtapaChangeView(LoginRequiredMixin, ContactoQuerysetMixin, View):
    def post(self, request, pk):
        contacto = get_object_or_404(self.get_base_queryset(), pk=pk)
        form = EtapaChangeForm(request.POST, contacto=contacto)
        if form.is_valid():
            nueva = form.cleaned_data['etapa']

            # Stage gate: verificar campos requeridos antes de avanzar
            campos = list(CampoPersonalizado.objects.prefetch_related('reglas').filter(
                activo=True
            ).filter(Q(tipo_contacto__isnull=True) | Q(tipo_contacto=contacto.tipo)))
            bloqueantes = []
            for campo in campos:
                for regla in campo.reglas.all():
                    if (regla.accion == 'obligatorio'
                            and regla.condicion_tipo == 'etapa'
                            and regla.condicion_etapa_id == nueva.pk):
                        val = (contacto.datos_extra or {}).get(campo.slug)
                        if not val:
                            bloqueantes.append(campo.nombre)
            if bloqueantes:
                lista = ', '.join(f'"{b}"' for b in bloqueantes)
                messages.error(request, f'Para avanzar a "{nueva.nombre}" completá: {lista}.')
                return redirect('contactos:detail', pk=pk)

            etapa_anterior = contacto.stage
            contacto.stage = nueva
            contacto.stage_updated_at = timezone.now()
            motivo = form.cleaned_data.get('motivo_perdida', '').strip()
            update_fields = ['stage', 'stage_updated_at', 'updated_at']
            if motivo:
                contacto.motivo_perdida = motivo
                update_fields.append('motivo_perdida')
            contacto.save(update_fields=update_fields)
            HistorialEtapa.objects.create(
                contacto=contacto,
                etapa_anterior=etapa_anterior,
                etapa_nueva=nueva,
                cambiado_por=request.user,
                nota=form.cleaned_data.get('nota', ''),
            )
            messages.success(request, f'Etapa cambiada a "{nueva.nombre}".')
            try:
                from apps.automations.tasks import iniciar_flujos_para_contacto
                from apps.automations.models import Flujo
                iniciar_flujos_para_contacto(contacto, Flujo.TRIGGER_ETAPA_CAMBIA)
            except Exception:
                pass
        else:
            for field, errs in form.errors.items():
                for err in errs:
                    messages.error(request, err)
        return redirect('contactos:detail', pk=pk)


class ContactoKanbanMoveView(LoginRequiredMixin, ContactoQuerysetMixin, View):
    """AJAX endpoint for Kanban drag & drop."""

    def post(self, request, pk):
        contacto = get_object_or_404(self.get_base_queryset(), pk=pk)
        etapa_id = request.POST.get('etapa')
        try:
            nueva = contacto.tipo.pipeline.stages.get(pk=etapa_id) if contacto.tipo.pipeline_id else None
        except (PipelineStage.DoesNotExist, ValueError, TypeError):
            nueva = None
        if not nueva:
            return JsonResponse({'error': 'Etapa inválida.'}, status=400)
        etapa_anterior = contacto.stage
        contacto.stage = nueva
        contacto.stage_updated_at = timezone.now()
        contacto.save(update_fields=['stage', 'stage_updated_at', 'updated_at'])
        HistorialEtapa.objects.create(
            contacto=contacto,
            etapa_anterior=etapa_anterior,
            etapa_nueva=nueva,
            cambiado_por=request.user,
        )
        return JsonResponse({'ok': True, 'etapa': nueva.pk, 'etapa_nombre': nueva.nombre})


class ContactSearchAPIView(LoginRequiredMixin, View):
    """AJAX: search contacts for quick lookup."""

    def get(self, request):
        q = request.GET.get('q', '').strip()
        qs = Contacto.objects.filter(telefono__startswith='+').order_by('nombre_completo')
        if q:
            qs = qs.filter(
                Q(nombre_completo__icontains=q) |
                Q(telefono__icontains=q) |
                Q(email__icontains=q)
            )
        results = [
            {
                'id': c.pk,
                'nombre': c.nombre_completo,
                'telefono': c.telefono,
                'email': c.email or '',
                'tipo': c.tipo.slug,
            }
            for c in qs[:20]
        ]
        return JsonResponse({'results': results})


class ContactoCSVImportView(LoginRequiredMixin, UserPassesTestMixin, View):
    template_name = 'contactos/contacto_import.html'

    def test_func(self):
        return self.request.user.can_see_all_leads

    def get(self, request):
        return render(request, self.template_name, {'form': ContactoImportForm()})

    def post(self, request):
        form = ContactoImportForm(request.POST, request.FILES)
        if not form.is_valid():
            return render(request, self.template_name, {'form': form})

        archivo = request.FILES['archivo']
        default_tipo = form.cleaned_data['tipo']
        actualizar = form.cleaned_data.get('actualizar_existentes', True)
        asignar = form.cleaned_data.get('asignar_agente', False)
        default_agente = request.user if asignar else None

        try:
            headers, rows = _read_file(archivo)
        except Exception as e:
            messages.error(request, f'Error leyendo el archivo: {e}')
            return render(request, self.template_name, {'form': form})

        if not rows:
            messages.warning(request, 'El archivo está vacío o no tiene filas de datos.')
            return render(request, self.template_name, {'form': form})

        plans_by_name = {p.nombre.lower(): p for p in Plan.objects.filter(activo=True)}
        tipos_by_slug = {t.slug: t for t in TipoContacto.objects.filter(activo=True)}
        tipos_by_slug.update({t.nombre.lower(): t for t in TipoContacto.objects.filter(activo=True)})

        agentes_cache = {}
        for u in User.objects.filter(is_active=True):
            agentes_cache[u.get_full_name().lower()] = u
            agentes_cache[u.username.lower()] = u
            agentes_cache[u.email.lower()] = u

        stats = {'created': 0, 'updated': 0, 'skipped': 0, 'error': 0}
        errors = []
        all_extra_keys = set()

        for i, row in enumerate(rows, start=2):
            action, msg, extra_keys = _process_row(row, plans_by_name, tipos_by_slug, default_tipo,
                                                    actualizar, default_agente, agentes_cache)
            stats[action] += 1
            all_extra_keys.update(extra_keys)
            if action == 'error':
                errors.append({'fila': i, 'msg': msg, 'datos': str(row)[:120]})

        ctx = {
            'form': ContactoImportForm(),
            'resultado': True,
            'stats': stats,
            'errors': errors[:50],
            'total_filas': len(rows),
            'extra_keys': sorted(all_extra_keys),
        }
        return render(request, self.template_name, ctx)


class ContactoCSVExportView(LoginRequiredMixin, ContactoQuerysetMixin, View):
    def get(self, request):
        qs = self.get_base_queryset()
        extra_keys = set()
        for contacto in qs:
            extra_keys.update((contacto.datos_extra or {}).keys())
        extra_keys = sorted(extra_keys)

        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="contactos.csv"'
        response.write('﻿')
        writer = csv.writer(response)
        base_headers = ['ID', 'tipo', 'nombre_completo', 'dni', 'telefono', 'email', 'localidad', 'provincia',
                        'plan', 'etapa', 'prioridad', 'origen', 'notas', 'agente', 'creado']
        writer.writerow(base_headers + extra_keys)
        for contacto in qs:
            base = [
                contacto.pk, contacto.tipo.nombre, contacto.nombre_completo, contacto.dni, contacto.telefono, contacto.email,
                contacto.localidad, contacto.provincia,
                contacto.plan_interes.nombre if contacto.plan_interes else '',
                contacto.stage.nombre if contacto.stage else '',
                contacto.get_prioridad_display(), contacto.get_origen_display(),
                contacto.notas,
                contacto.agente.get_full_name() if contacto.agente else '',
                contacto.created_at.strftime('%d/%m/%Y %H:%M'),
            ]
            extras = [(contacto.datos_extra or {}).get(k, '') for k in extra_keys]
            writer.writerow(base + extras)
        return response


class ContactoImportTemplateView(LoginRequiredMixin, View):
    """Download an Excel import template with column explanations."""

    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Contactos'

        headers = [
            'nombre_completo', 'codigo_pais', 'telefono', 'email', 'dni',
            'localidad', 'provincia', 'tipo', 'plan', 'notas', 'origen',
            'empresa', 'cargo',
        ]
        descriptions = [
            'Nombre y apellido *', 'Cód. de país (ej: 54)', 'Teléfono sin código de país *',
            'Email', 'DNI (7-8 dígitos)', 'Localidad / Ciudad', 'Provincia',
            'lead / cliente / (otro tipo configurado)',
            'Nombre exacto del plan', 'Notas internas',
            'web / campaña / referido / llamada / whatsapp',
            'Columna extra de ejemplo', 'Otra columna extra',
        ]

        header_fill = PatternFill('solid', fgColor='1E3A5F')
        desc_fill = PatternFill('solid', fgColor='E8F0FE')
        extra_fill = PatternFill('solid', fgColor='FFF3CD')
        bold_white = Font(bold=True, color='FFFFFF')

        for col, (header, desc) in enumerate(zip(headers, descriptions), start=1):
            is_extra = col > 11
            hcell = ws.cell(row=1, column=col, value=header)
            hcell.font = bold_white
            hcell.fill = extra_fill if is_extra else header_fill
            hcell.alignment = Alignment(horizontal='center')
            dcell = ws.cell(row=2, column=col, value=desc)
            dcell.fill = PatternFill('solid', fgColor='FFF3CD') if is_extra else desc_fill
            dcell.font = Font(italic=True, color='666666' if not is_extra else '856404')
            ws.column_dimensions[get_column_letter(col)].width = max(len(header), len(desc)) + 4

        examples = [
            ['Juan Pérez', '54', '1123456789', 'juan@gmail.com', '35123456', 'Buenos Aires', 'Buenos Aires', 'lead', 'Plan Individual', '', 'web', 'Empresa SA', 'Gerente'],
            ['María García', '54', '2996543210', 'maria@gmail.com', '', 'Neuquén', 'Neuquén', 'cliente', '', 'Interesada en familiar', 'whatsapp', '', ''],
        ]
        for r, row in enumerate(examples, start=3):
            for c, val in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=val)

        ws2 = wb.create_sheet('Instrucciones')
        instructions = [
            ('INSTRUCCIONES DE IMPORTACIÓN', True),
            ('', False),
            ('COLUMNAS REQUERIDAS (* obligatorias):', True),
            ('  nombre_completo — Nombre y apellido del contacto', False),
            ('  telefono — Solo el número, sin código de país ni espacios ni guiones', False),
            ('', False),
            ('CÓDIGO DE PAÍS:', True),
            ('  codigo_pais — Ej: 54 (Argentina), 598 (Uruguay), 56 (Chile)', False),
            ('  Si no se incluye esta columna, se asume 54 (Argentina)', False),
            ('', False),
            ('COLUMNAS ESTÁNDAR (opcionales):', True),
            ('  email, dni, localidad, provincia, tipo, plan, notas, origen', False),
            ('  tipo: clave del tipo de contacto (ej: lead, cliente). Si se omite, se usa el seleccionado en el formulario.', False),
            ('  origen válidos: web, campaña, referido, llamada, whatsapp', False),
            ('', False),
            ('COLUMNAS EXTRAS (datos personalizados):', True),
            ('  Podés agregar cualquier otra columna con el nombre que quieras.', False),
            ('  Ej: empresa, cargo, fecha_cumpleaños, numero_socio', False),
            ('  Estos datos se guardan en el contacto y pueden usarse como variables', False),
            ('  en las plantillas de WhatsApp al crear campañas masivas.', False),
            ('', False),
            ('COMPORTAMIENTO DEL IMPORT:', True),
            ('  - Si un número de teléfono ya existe → se actualiza el contacto', False),
            ('  - Si el número no existe → se crea un contacto nuevo', False),
            ('  - Teléfonos duplicados dentro del archivo → se procesa el primero', False),
            ('', False),
            ('FORMATOS DE TELÉFONO ACEPTADOS:', True),
            ('  1123456789  →  +541123456789', False),
            ('  011 2345-6789  →  +541123456789', False),
            ('  +54 11 2345-6789  →  +541123456789', False),
        ]
        for row_idx, (text, is_bold) in enumerate(instructions, start=1):
            cell = ws2.cell(row=row_idx, column=1, value=text)
            if is_bold:
                cell.font = Font(bold=True)
        ws2.column_dimensions['A'].width = 65

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="template_importar_contactos.xlsx"'
        wb.save(response)
        return response


class ContactoBulkAssignView(LoginRequiredMixin, UserPassesTestMixin, View):
    def test_func(self):
        return self.request.user.can_see_all_leads

    def post(self, request):
        contacto_ids = request.POST.getlist('contacto_ids') or request.POST.getlist('lead_ids')
        agente_id = request.POST.get('agente_id', '').strip()

        if not contacto_ids:
            messages.warning(request, 'No seleccionaste ningún contacto.')
            return redirect(request.POST.get('next', 'contactos:list'))

        try:
            agente = User.objects.get(pk=agente_id, is_active=True)
        except (User.DoesNotExist, ValueError):
            messages.error(request, 'Agente inválido.')
            return redirect(request.POST.get('next', 'contactos:list'))

        updated = Contacto.objects.filter(pk__in=contacto_ids).update(agente=agente)
        messages.success(request, f'{updated} contacto{"s" if updated != 1 else ""} asignado{"s" if updated != 1 else ""} a {agente.get_full_name()}.')
        return redirect(request.POST.get('next', 'contactos:list'))


# ── Notas internas ────────────────────────────────────────

class NotaContactoCreateView(LoginRequiredMixin, ContactoQuerysetMixin, View):
    def post(self, request, pk):
        contacto = get_object_or_404(self.get_base_queryset(), pk=pk)
        texto = request.POST.get('texto', '').strip()
        if not texto:
            return JsonResponse({'error': 'El texto no puede estar vacío.'}, status=400)
        nota = NotaContacto.objects.create(contacto=contacto, autor=request.user, texto=texto)
        return JsonResponse({
            'ok': True,
            'id': nota.pk,
            'texto': nota.texto,
            'autor': nota.autor.display_name,
            'created_at': nota.created_at.strftime('%d/%m/%Y %H:%M'),
        })


class NotaContactoDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        nota = get_object_or_404(NotaContacto, pk=pk)
        if nota.autor != request.user and not request.user.can_see_all_leads:
            return JsonResponse({'error': 'Sin permiso.'}, status=403)
        nota.delete()
        return JsonResponse({'ok': True})


# ── Etiquetas ─────────────────────────────────────────────

class EtiquetaToggleView(LoginRequiredMixin, ContactoQuerysetMixin, View):
    """AJAX: add or remove an etiqueta from a contacto."""
    def post(self, request, pk):
        contacto = get_object_or_404(self.get_base_queryset(), pk=pk)
        slug = request.POST.get('slug', '').strip()
        etiqueta = get_object_or_404(Etiqueta, slug=slug)
        if contacto.etiquetas.filter(pk=etiqueta.pk).exists():
            contacto.etiquetas.remove(etiqueta)
            accion = 'removed'
        else:
            contacto.etiquetas.add(etiqueta)
            accion = 'added'
        return JsonResponse({'ok': True, 'accion': accion, 'slug': etiqueta.slug, 'nombre': etiqueta.nombre, 'color': etiqueta.color})


class EtiquetaListView(LoginRequiredMixin, SupervisorRequiredMixin, View):
    template_name = 'contactos/etiqueta_list.html'

    def get(self, request):
        return render(request, self.template_name, {'etiquetas': Etiqueta.objects.all()})

    def post(self, request):
        nombre = request.POST.get('nombre', '').strip()
        color = request.POST.get('color', 'secondary')
        if not nombre:
            messages.error(request, 'El nombre es obligatorio.')
        elif Etiqueta.objects.filter(nombre__iexact=nombre).exists():
            messages.error(request, f'Ya existe una etiqueta llamada "{nombre}".')
        else:
            Etiqueta.objects.create(nombre=nombre, color=color)
            messages.success(request, f'Etiqueta "{nombre}" creada.')
        return redirect('contactos:etiquetas')


class EtiquetaDeleteView(LoginRequiredMixin, SupervisorRequiredMixin, View):
    def post(self, request, pk):
        etiqueta = get_object_or_404(Etiqueta, pk=pk)
        etiqueta.delete()
        messages.success(request, f'Etiqueta "{etiqueta.nombre}" eliminada.')
        return redirect('contactos:etiquetas')


# ── Filtros guardados ─────────────────────────────────────

class FiltroGuardadoCreateView(LoginRequiredMixin, View):
    def post(self, request):
        nombre = request.POST.get('nombre', '').strip()
        tipo_slug = request.POST.get('tipo', '').strip()
        if not nombre:
            return JsonResponse({'error': 'Nombre requerido.'}, status=400)
        tipo = TipoContacto.objects.filter(slug=tipo_slug).first() if tipo_slug else None
        # Capture all GET-style filter params from POST body
        params = {k: v for k, v in request.POST.items()
                  if k not in ('csrfmiddlewaretoken', 'nombre', 'tipo') and v}
        filtro, created = FiltroGuardado.objects.update_or_create(
            usuario=request.user, nombre=nombre,
            defaults={'tipo_contacto': tipo, 'parametros': params},
        )
        return JsonResponse({'ok': True, 'id': filtro.pk, 'nombre': filtro.nombre, 'created': created})


class FiltroGuardadoDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        filtro = get_object_or_404(FiltroGuardado, pk=pk, usuario=request.user)
        filtro.delete()
        return JsonResponse({'ok': True})


# ── Edición inline de campos ──────────────────────────────

_INLINE_EDITABLE = {
    'nombre_completo', 'dni', 'email', 'localidad', 'provincia',
    'fecha_nacimiento', 'telefono', 'grupo_familiar', 'prioridad', 'origen',
    'agente', 'plan_interes',
}

class ContactoFieldUpdateView(LoginRequiredMixin, ContactoQuerysetMixin, View):
    """AJAX PATCH: update a single standard field of a contacto inline."""
    def post(self, request, pk):
        contacto = get_object_or_404(self.get_base_queryset(), pk=pk)
        campo = request.POST.get('campo', '').strip()
        valor = request.POST.get('valor', '').strip()

        if campo not in _INLINE_EDITABLE:
            return JsonResponse({'error': 'Campo no editable.'}, status=400)

        # Special handling for FK / choice fields
        if campo == 'plan_interes':
            if valor:
                plan = Plan.objects.filter(pk=valor).first()
                if not plan:
                    return JsonResponse({'error': 'Plan inválido.'}, status=400)
                contacto.plan_interes = plan
            else:
                contacto.plan_interes = None
            contacto.save(update_fields=['plan_interes', 'updated_at'])
            return JsonResponse({'ok': True, 'display': str(contacto.plan_interes) if contacto.plan_interes else '—'})

        if campo == 'agente':
            if valor:
                from apps.users.models import User as UserModel
                agente = UserModel.objects.filter(pk=valor, is_active=True).first()
                if not agente:
                    return JsonResponse({'error': 'Agente inválido.'}, status=400)
                contacto.agente = agente
            else:
                contacto.agente = None
            contacto.save(update_fields=['agente', 'updated_at'])
            return JsonResponse({'ok': True, 'display': contacto.agente.display_name if contacto.agente else 'Sin asignar'})

        try:
            setattr(contacto, campo, valor or None if campo == 'fecha_nacimiento' else valor)
            contacto.full_clean(exclude=[f for f in _INLINE_EDITABLE if f != campo])
            contacto.save(update_fields=[campo, 'updated_at'])
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

        display = getattr(contacto, campo, '') or '—'
        return JsonResponse({'ok': True, 'display': str(display)})


# ── Detección de duplicados ───────────────────────────────

class ContactoDuplicateCheckView(LoginRequiredMixin, View):
    """AJAX: check if a phone or email is already in use."""
    def get(self, request):
        telefono = request.GET.get('telefono', '').strip()
        email = request.GET.get('email', '').strip()
        exclude_pk = request.GET.get('exclude', '').strip()

        duplicados = []
        qs = Contacto.objects.select_related('tipo')
        if exclude_pk:
            qs = qs.exclude(pk=exclude_pk)

        if telefono:
            for c in qs.filter(telefono=telefono)[:3]:
                duplicados.append({'pk': c.pk, 'nombre': c.nombre_completo, 'tipo': c.tipo.nombre, 'campo': 'teléfono'})
        if email:
            for c in qs.filter(email__iexact=email).exclude(pk__in=[d['pk'] for d in duplicados])[:3]:
                duplicados.append({'pk': c.pk, 'nombre': c.nombre_completo, 'tipo': c.tipo.nombre, 'campo': 'email'})

        return JsonResponse({'duplicados': duplicados})


# ── Búsqueda global ───────────────────────────────────────

class BusquedaGlobalView(LoginRequiredMixin, View):
    template_name = 'contactos/busqueda_global.html'

    def get(self, request):
        q = request.GET.get('q', '').strip()
        if len(q) < 2:
            return render(request, self.template_name, {'q': q, 'resultados': None})

        filtro_q = Q(nombre_completo__icontains=q) | Q(dni__icontains=q) | Q(telefono__icontains=q) | Q(email__icontains=q)
        contactos_qs = Contacto.objects.select_related('tipo', 'stage', 'agente').filter(filtro_q)
        if not request.user.can_see_all_leads:
            contactos_qs = contactos_qs.filter(agente=request.user)

        from apps.tasks.models import Tarea
        tareas_qs = Tarea.objects.select_related('contacto').filter(
            Q(descripcion__icontains=q) | Q(contacto__nombre_completo__icontains=q)
        )
        if not request.user.can_see_all_leads:
            tareas_qs = tareas_qs.filter(agente=request.user)

        from apps.whatsapp.models import Conversacion
        convs_qs = Conversacion.objects.filter(
            Q(nombre_contacto__icontains=q) | Q(telefono__icontains=q)
        )
        if not request.user.can_see_all_leads:
            convs_qs = convs_qs.filter(agente=request.user)

        return render(request, self.template_name, {
            'q': q,
            'resultados': {
                'contactos': contactos_qs[:10],
                'tareas':    tareas_qs[:5],
                'convs':     convs_qs[:5],
            },
        })
